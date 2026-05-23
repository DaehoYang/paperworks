#!/usr/bin/env python3
"""Generate meeting minutes PDFs from receipt metadata."""

from __future__ import annotations

import argparse
import csv
import json
import math
import os
import re
import subprocess
import tempfile
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Iterable

from PIL import Image, ImageOps
from pypdf import PdfReader, PdfWriter
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfgen import canvas

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

try:
    import yaml
except ImportError:  # pragma: no cover
    yaml = None


BASE_DIR = Path(__file__).resolve().parent
FORMAT_DIR = BASE_DIR / "format"
RECEIPT_DIR = BASE_DIR / "receipt"
OUTPUT_DIR = BASE_DIR / "output"

TEMPLATE_PDF = FORMAT_DIR / "바나연회의록_빈칸.pdf"
INFO_YML = FORMAT_DIR / "information.yml"
INFO_XLSX = FORMAT_DIR / "information.xlsx"
MEMBERS_XLSX = FORMAT_DIR / "members.xlsx"
SUMMARY_CSV = RECEIPT_DIR / "summary.csv"

SUMMARY_HEADER = [
    "file_name",
    "total_price",
    "store_name",
    "address",
    "meeting_place",
    "generated",
    "topic",
    "attendee_count",
    "item_count",
    "food_count",
    "drink_count",
    "attendee_names",
    "ocr_engine",
]
SEOUL_WEST_DISTRICTS = ("강서구", "양천구", "마포구", "서대문구", "은평구")
TOPIC_ORDER = ("quantum", "holography", "deeplearning")

TOPICS = {
    "quantum": (
        "양자광학 정기 랩미팅",
        [
            "최근 양자광학 연구 진행 상황 공유",
            "앞으로의 연구 방향 논의",
            "논문 작성 검토",
        ],
    ),
    "holography": (
        "홀로그래피 정기 랩미팅",
        [
            "최근 홀로그래피 연구 진행 상황 공유",
            "앞으로의 연구 방향 논의",
            "논문 작성 검토",
        ],
    ),
    "deeplearning": (
        "딥러닝 연구 정기 랩미팅",
        [
            "최근 딥러닝 응용 연구 진행 상황 공유",
            "앞으로의 연구 방향 논의",
            "논문 작성 검토",
        ],
    ),
}


@dataclass(frozen=True)
class Member:
    department: str
    position: str
    name: str


@dataclass(frozen=True)
class ReceiptMetadata:
    receipt_path: Path
    total_price: int
    generated: datetime
    store_name: str
    address: str
    topic: str
    item_count: int | None = None
    food_count: int | None = None
    drink_count: int | None = None
    ocr_engine: str = "manual"

    @property
    def receipt_name(self) -> str:
        return self.receipt_path.name

    @property
    def stem(self) -> str:
        return self.receipt_path.stem


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Fill the meeting minutes PDF and attach a receipt image."
    )
    parser.add_argument("receipt", help="Receipt image path, relative to meeting/receipt or absolute.")
    parser.add_argument("--total-price", type=int, help="Total receipt amount in KRW.")
    parser.add_argument(
        "--generated",
        help="Receipt date/time. Use 'YYYY-MM-DD HH:MM:SS' or 'YYYY-MM-DD HH:MM'.",
    )
    parser.add_argument("--store-name", help="Store/cafe/restaurant name from the receipt.")
    parser.add_argument("--address", help="Address from the receipt.")
    parser.add_argument(
        "--topic",
        default="auto",
        help="Meeting content preset. Use auto to rotate topics from summary.csv.",
    )
    parser.add_argument("--item-count", type=int, help="Number of food/drink items inferred from receipt.")
    parser.add_argument("--food-count", type=int, help="Number of meal/main food items inferred from receipt.")
    parser.add_argument("--drink-count", type=int, help="Number of drink items inferred from receipt.")
    parser.add_argument(
        "--no-summary",
        action="store_true",
        help="Do not update receipt/summary.csv.",
    )
    parser.add_argument(
        "--ocr",
        action="store_true",
        help="Read total-price/generated/store-name/address automatically.",
    )
    parser.add_argument(
        "--ocr-engine",
        choices=["codex", "paddle-litellm"],
        default="codex",
        help="OCR backend used when --ocr is set.",
    )
    parser.add_argument("--codex-bin", default="codex", help="Codex CLI executable.")
    parser.add_argument("--ocr-model", help="Optional Codex model for OCR.")
    parser.add_argument("--ocr-timeout", type=int, default=180, help="OCR timeout in seconds.")
    parser.add_argument(
        "--litellm-base-url",
        default=os.environ.get("DHLAB_LITELLM_BASE_URL", "https://dhlab.gachon.ac.kr/services/litellm/v1"),
        help="OpenAI-compatible LiteLLM base URL for paddle-litellm OCR.",
    )
    parser.add_argument(
        "--litellm-api-key",
        default=os.environ.get("DHLAB_LITELLM_API_KEY"),
        help="LiteLLM API key. Defaults to DHLAB_LITELLM_API_KEY.",
    )
    parser.add_argument(
        "--litellm-model",
        default=os.environ.get("DHLAB_LITELLM_MODEL", "local"),
        help="LiteLLM model name.",
    )
    return parser.parse_args()


def parse_datetime(raw: str) -> datetime:
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            pass
    raise ValueError(f"unsupported datetime format: {raw}")


def maybe_fix_date_from_filename(receipt_path: Path, generated: datetime) -> datetime:
    match = re.search(r"(20\d{2})(\d{2})(\d{2})", receipt_path.name)
    if not match:
        return generated
    year, month, day = map(int, match.groups())
    if generated.month == month and generated.day == day and generated.year != year:
        return generated.replace(year=year)
    return generated


def resolve_receipt_path(raw: str) -> Path:
    path = Path(raw)
    if not path.is_absolute():
        path = RECEIPT_DIR / path
    path = path.resolve()
    if not path.exists():
        raise FileNotFoundError(path)
    return path


def extract_json_object(raw: str) -> dict[str, object]:
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", raw, flags=re.DOTALL)
        if not match:
            raise ValueError(f"Codex OCR did not return JSON: {raw!r}") from None
        parsed = json.loads(match.group(0))
    if not isinstance(parsed, dict):
        raise ValueError(f"Codex OCR returned non-object JSON: {parsed!r}")
    return parsed


def run_codex_ocr(
    receipt_path: Path,
    codex_bin: str = "codex",
    model: str | None = None,
    timeout: int = 180,
) -> dict[str, object]:
    prompt = (
        "Read this Korean receipt image. Return only compact JSON with exactly these keys: "
        "total_price integer, generated string in YYYY-MM-DD HH:MM:SS, store_name string, "
        "address string. Use the receipt date printed on the image. Do not include markdown."
    )
    with tempfile.NamedTemporaryFile("w+", suffix=".json", delete=False) as handle:
        output_path = Path(handle.name)

    command = [
        codex_bin,
        "exec",
        "--skip-git-repo-check",
        "--ephemeral",
        "-s",
        "read-only",
        "--color",
        "never",
        "--image",
        str(receipt_path),
        "-o",
        str(output_path),
    ]
    if model:
        command.extend(["--model", model])
    command.append(prompt)

    try:
        completed = subprocess.run(
            command,
            input="",
            text=True,
            capture_output=True,
            timeout=timeout,
            check=False,
        )
        if completed.returncode != 0:
            raise RuntimeError(
                "Codex OCR failed with exit code "
                f"{completed.returncode}\nSTDERR:\n{completed.stderr[-4000:]}"
            )
        raw = output_path.read_text(encoding="utf-8").strip()
        return extract_json_object(raw)
    finally:
        output_path.unlink(missing_ok=True)


def run_paddle_ocr_text(receipt_path: Path) -> str:
    # Import lazily because PaddleOCR is heavy and optional for the Codex path.
    try:
        import pyclipper  # noqa: F401
        from paddleocr import PaddleOCR
    except Exception as exc:  # pragma: no cover
        raise RuntimeError(
            "PaddleOCR dependencies are not available. Install paddlepaddle and paddleocr."
        ) from exc

    ocr = PaddleOCR(lang="korean", use_angle_cls=True, show_log=False, use_gpu=False)
    image = ImageOps.exif_transpose(Image.open(receipt_path)).convert("RGB")
    with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as handle:
        ocr_input = Path(handle.name)
    try:
        image.save(ocr_input, quality=95)
        result = ocr.ocr(str(ocr_input), cls=True)
    finally:
        ocr_input.unlink(missing_ok=True)
    lines: list[tuple[float, float, str, float]] = []
    for page in result or []:
        for item in page or []:
            try:
                box, (text, confidence) = item
                y = sum(point[1] for point in box) / len(box)
                x = sum(point[0] for point in box) / len(box)
                lines.append((y, x, str(text), float(confidence)))
            except Exception:
                continue

    lines.sort(key=lambda row: (row[0], row[1]))
    return "\n".join(f"{text}\t(conf={confidence:.3f})" for _y, _x, text, confidence in lines)


def normalize_amount(raw: str) -> int | None:
    digits = re.sub(r"\D", "", raw)
    if not digits:
        return None
    value = int(digits)
    if value < 1000 or value > 1000000:
        return None
    return value


def ocr_hints(ocr_text: str) -> dict[str, object]:
    plain_lines = [line.split("\t", 1)[0].strip() for line in ocr_text.splitlines()]
    plain_lines = [line for line in plain_lines if line]
    plain = "\n".join(plain_lines)

    amount_counts: dict[int, int] = {}
    for raw in re.findall(r"(?<!\d)\d[\d,\s]{2,10}(?!\d)", plain):
        amount = normalize_amount(raw)
        if amount is None:
            continue
        # Exclude obvious phone/date/business-number fragments.
        if amount < 5000:
            continue
        amount_counts[amount] = amount_counts.get(amount, 0) + 1
    amount_candidates = [
        {"amount": amount, "count": count}
        for amount, count in sorted(amount_counts.items(), key=lambda item: (-item[1], -item[0]))[:20]
    ]

    label_windows: list[str] = []
    labels = ("합", "계", "총액", "주문", "카드", "금액", "공급대가", "결제", "신용")
    for idx, line in enumerate(plain_lines):
        if any(label in line for label in labels):
            start = max(0, idx - 3)
            end = min(len(plain_lines), idx + 8)
            label_windows.append(" | ".join(plain_lines[start:end]))
    label_windows = label_windows[:20]

    date_candidates = re.findall(
        r"20\d{2}[./년 -]?\s*\d{1,2}[./월 -]?\s*\d{1,2}[일]?(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?",
        plain,
    )
    address_candidates = [
        line
        for line in plain_lines
        if any(token in line for token in ("서울", "경기", "성남", "강서구", "성동구", "수정구", "마곡", "왕십리", "로", "길", "동"))
    ][:30]

    return {
        "amount_candidates_by_frequency": amount_candidates,
        "label_windows": label_windows,
        "date_candidates": date_candidates[:10],
        "address_candidates": address_candidates,
        "first_lines": plain_lines[:25],
    }


def run_litellm_receipt_parser(
    ocr_text: str,
    base_url: str,
    api_key: str,
    model: str = "local",
    timeout: int = 120,
) -> dict[str, object]:
    import urllib.request

    url = base_url.rstrip("/") + "/chat/completions"
    prompt = (
        "Extract Korean receipt fields from OCR text. Output only compact JSON with exactly "
        "these keys: total_price, generated, store_name, address, item_count, food_count, drink_count. total_price must be an "
        "integer KRW. generated must be YYYY-MM-DD HH:MM:SS. Pick the actual receipt total, "
        "not a business registration number, phone number, approval number, tax-only amount, "
        "or supply-only amount. Prefer repeated amount candidates and labels like 합계, 총액, "
        "주문총액, 카드금액, 결제금액, 공급대가. Reconstruct addresses from OCR fragments. "
        "item_count is the total quantity of ordered food/drink items, excluding discounts, tax, "
        "card/payment lines, approval lines, and receipt metadata. food_count counts meals/main dishes; "
        "drink_count counts beverages. If a restaurant receipt has meals plus drinks, do not treat drinks "
        "as extra people by themselves. If quantities are unclear, infer conservatively. Do not add extra keys."
    )
    hints = ocr_hints(ocr_text)
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": prompt},
            {
                "role": "user",
                "content": (
                    "Heuristic OCR hints:\n"
                    f"{json.dumps(hints, ensure_ascii=False)}\n\n"
                    f"Raw OCR text:\n{ocr_text[:12000]}"
                ),
            },
        ],
        "temperature": 0,
        "max_tokens": 500,
        "extra_body": {"chat_template_kwargs": {"enable_thinking": False}},
    }
    request = urllib.request.Request(
        url,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
    )
    with urllib.request.urlopen(request, timeout=timeout) as response:
        data = json.loads(response.read().decode("utf-8"))

    message = data["choices"][0]["message"]
    content = message.get("content")
    if not content:
        content = message.get("reasoning_content") or ""
    return extract_json_object(content)


def run_paddle_litellm_ocr(
    receipt_path: Path,
    base_url: str,
    api_key: str,
    model: str,
    timeout: int,
) -> dict[str, object]:
    if not api_key:
        raise ValueError("--litellm-api-key or DHLAB_LITELLM_API_KEY is required for paddle-litellm")
    ocr_text = run_paddle_ocr_text(receipt_path)
    parsed = run_litellm_receipt_parser(
        ocr_text=ocr_text,
        base_url=base_url,
        api_key=api_key,
        model=model,
        timeout=timeout,
    )
    parsed["_ocr_text"] = ocr_text
    return parsed


def require_value(value: object, name: str) -> object:
    if value is None or value == "":
        raise ValueError(f"{name} is required unless --ocr can read it")
    return value


def metadata_from_args(args: argparse.Namespace) -> ReceiptMetadata:
    receipt_path = resolve_receipt_path(args.receipt)
    ocr_data: dict[str, object] = {}
    if args.ocr:
        if args.ocr_engine == "codex":
            ocr_data = run_codex_ocr(
                receipt_path,
                codex_bin=args.codex_bin,
                model=args.ocr_model,
                timeout=args.ocr_timeout,
            )
        elif args.ocr_engine == "paddle-litellm":
            ocr_data = run_paddle_litellm_ocr(
                receipt_path,
                base_url=args.litellm_base_url,
                api_key=args.litellm_api_key,
                model=args.litellm_model,
                timeout=args.ocr_timeout,
            )
        else:  # pragma: no cover
            raise ValueError(f"unsupported OCR engine: {args.ocr_engine}")

    total_price = args.total_price
    if total_price is None and "total_price" in ocr_data:
        total_price = int(str(ocr_data["total_price"]).replace(",", ""))

    generated_raw = args.generated or ocr_data.get("generated")
    generated = parse_datetime(str(require_value(generated_raw, "generated")))
    generated = maybe_fix_date_from_filename(receipt_path, generated)

    return ReceiptMetadata(
        receipt_path=receipt_path,
        total_price=int(require_value(total_price, "total_price")),
        generated=generated,
        store_name=str(require_value(args.store_name or ocr_data.get("store_name"), "store_name")),
        address=str(require_value(args.address or ocr_data.get("address"), "address")),
        topic=args.topic,
        item_count=args.item_count or safe_int(ocr_data.get("item_count")),
        food_count=args.food_count or safe_int(ocr_data.get("food_count")),
        drink_count=args.drink_count or safe_int(ocr_data.get("drink_count")),
        ocr_engine=args.ocr_engine if args.ocr else "manual",
    )


def safe_int(value: object) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(str(value).replace(",", ""))
    except ValueError:
        return None


def derive_meeting_place(address: str, store_name: str) -> str:
    places = configured_meeting_places()
    gachon_districts = configured_districts("gachon_univ_districts", ("성남시", "수정구"))
    ewha_mokdong_districts = configured_districts("ewha_mokdong_districts", ("양천구", "목동", "영등포구"))
    seoul_west_districts = configured_districts("seoul_west_districts", SEOUL_WEST_DISTRICTS)
    kriss_districts = configured_districts("kriss_districts", ("대전광역시", "대전시", "유성구", "도룡동"))
    if any(district in address for district in gachon_districts):
        return places["gachon_univ"]
    if any(district in address for district in ewha_mokdong_districts):
        return places["ewha_mokdong"]
    if any(district in address for district in seoul_west_districts):
        return places["seoul_west"]
    if any(district in address for district in kriss_districts):
        return places["kriss"]
    return store_name


def attendee_count(total_price: int) -> int:
    rules = configured_attendee_rules()
    price_per_person = int(rules.get("price_per_person") or 30000)
    return max(1, math.ceil(total_price / price_per_person))


def attendee_count_from_items(
    item_count: int | None,
    food_count: int | None,
    drink_count: int | None,
) -> int:
    food = max(0, food_count or 0)
    drink = max(0, drink_count or 0)
    total = max(item_count or 0, food + drink)
    if total <= 0:
        return 1
    if food == 0 and drink == 0:
        return total
    if food > 0 and drink > 0:
        return max(food, math.ceil(total / 2))
    if food > 0:
        return food
    return drink


def adjusted_attendee_count(
    total_price: int,
    item_count: int | None,
    food_count: int | None = None,
    drink_count: int | None = None,
) -> int:
    max_attendees = int(configured_attendee_rules().get("max_attendees") or 10)
    return min(
        max_attendees,
        max(attendee_count(total_price), attendee_count_from_items(item_count, food_count, drink_count)),
    )


def read_summary_records() -> list[dict[str, str]]:
    if not SUMMARY_CSV.exists():
        return []
    with SUMMARY_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    if not rows:
        return []
    header = normalize_summary_header(rows[0])
    records: list[dict[str, str]] = []
    for row in rows[1:]:
        if not row or not any(cell.strip() for cell in row):
            continue
        padded = row + [""] * (len(header) - len(row))
        records.append(dict(zip(header, padded)))
    return records


def normalize_summary_header(header: list[str]) -> list[str]:
    normalized = [column.strip() for column in header]
    if "file_name" not in normalized:
        return SUMMARY_HEADER
    if "meeting_place" not in normalized and "location" in normalized:
        normalized[normalized.index("location")] = "meeting_place"
    return normalized


def choose_auto_topic(receipt_name: str, *, external: bool = False) -> str:
    records = [record for record in read_summary_records() if record.get("file_name") != receipt_name]
    topic_order = configured_external_topic_order() if external else configured_topic_order()
    used_topics = [record.get("topic", "") for record in records if record.get("topic") in topic_order]
    if not used_topics:
        return topic_order[0]
    last_index = topic_order.index(used_topics[-1])
    return topic_order[(last_index + 1) % len(topic_order)]


def previously_used_names(receipt_name: str) -> set[str]:
    names: set[str] = set()
    for record in read_summary_records():
        if record.get("file_name") == receipt_name:
            continue
        for name in record.get("attendee_names", "").split(";"):
            name = name.strip()
            if name:
                names.add(name)
    return names


def select_attendees(
    members: list[Member],
    count: int,
    metadata: ReceiptMetadata,
    external_members: list[Member] | None = None,
) -> list[Member]:
    fixed_first = str(configured_attendee_rules().get("fixed_first_attendee") or "양대호")
    lead = next((member for member in members if member.name == fixed_first), None)
    external_members = external_members or []
    required = ([lead] if lead is not None else []) + external_members
    required_names = {member.name for member in required}
    count = max(count, len(required))
    if count >= len(members):
        if lead is None:
            return (external_members + members)[:count]
        return (required + [member for member in members if member.name not in required_names])[:count]
    used = previously_used_names(metadata.receipt_name)
    others = [member for member in members if member.name not in required_names]
    seed = int(metadata.generated.strftime("%Y%m%d")) + sum(ord(ch) for ch in metadata.receipt_name)
    if others:
        shift = seed % len(others)
        rotated = others[shift:] + others[:shift]
    else:
        rotated = []
    fresh = [member for member in rotated if member.name not in used]
    repeated = [member for member in rotated if member.name in used]
    selected = required + fresh + repeated
    deduped: list[Member] = []
    seen: set[str] = set()
    for member in selected:
        if member.name in seen:
            continue
        deduped.append(member)
        seen.add(member.name)
        if len(deduped) == count:
            return deduped
    for member in members:
        if member.name not in seen:
            deduped.append(member)
            seen.add(member.name)
        if len(deduped) == count:
            break
    return deduped


def read_project_info(path: Path = INFO_XLSX) -> dict[str, str]:
    if INFO_YML.exists():
        config = read_information_yml()
        project = config.get("project", {})
        return {str(key): str(value) for key, value in project.items() if value is not None}
    if load_workbook is None:
        return {
            "과제번호": "202511000001",
            "연구책임자": "박영서",
            "연구과제명": "2025-2 가천바이오나노연구원 운영비",
            "연구기간": "25-09-01 ~ 26-08-31",
        }
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    info: dict[str, str] = {}
    for field, value in sheet.iter_rows(min_row=2, values_only=True):
        if field is None or value is None:
            continue
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        info[str(field)] = str(value)
    return info


def read_information_yml(path: Path = INFO_YML) -> dict[str, object]:
    if yaml is None:
        raise RuntimeError("PyYAML is required to read meeting/format/information.yml")
    with path.open("r", encoding="utf-8") as handle:
        data = yaml.safe_load(handle) or {}
    if not isinstance(data, dict):
        raise ValueError(f"{path} must contain a YAML mapping")
    return data


def configured_topics() -> dict[str, tuple[str, list[str]]]:
    if not INFO_YML.exists():
        return TOPICS
    raw_topics = read_information_yml().get("topics", {})
    if not isinstance(raw_topics, dict) or not raw_topics:
        return TOPICS
    topics: dict[str, tuple[str, list[str]]] = {}
    for key, value in raw_topics.items():
        if not isinstance(value, dict):
            continue
        title = value.get("title")
        content = value.get("content", [])
        if not title or not isinstance(content, list):
            continue
        topics[str(key)] = (str(title), [str(line) for line in content])
    return topics or TOPICS


def configured_external_topics() -> dict[str, tuple[str, list[str]]]:
    if not INFO_YML.exists():
        return {}
    raw_topics = read_information_yml().get("external_topics", {})
    if not isinstance(raw_topics, dict):
        return {}
    topics: dict[str, tuple[str, list[str]]] = {}
    for key, value in raw_topics.items():
        if not isinstance(value, dict):
            continue
        title = value.get("title")
        content = value.get("content", [])
        if not title or not isinstance(content, list):
            continue
        topics[str(key)] = (str(title), [str(line) for line in content])
    return topics


def configured_all_topics() -> dict[str, tuple[str, list[str]]]:
    return {**configured_topics(), **configured_external_topics()}


def configured_topic_order() -> tuple[str, ...]:
    if not INFO_YML.exists():
        return TOPIC_ORDER
    config = read_information_yml()
    order = config.get("topic_order", [])
    topics = configured_topics()
    if isinstance(order, list):
        cleaned = tuple(str(key) for key in order if str(key) in topics)
        if cleaned:
            return cleaned
    return tuple(topics)


def configured_external_topic_order() -> tuple[str, ...]:
    topics = configured_external_topics()
    if not topics:
        return configured_topic_order()
    if not INFO_YML.exists():
        return tuple(topics)
    config = read_information_yml()
    order = config.get("external_topic_order", [])
    if isinstance(order, list):
        cleaned = tuple(str(key) for key in order if str(key) in topics)
        if cleaned:
            return cleaned
    return tuple(topics)


def configured_attendee_rules() -> dict[str, object]:
    defaults = {
        "price_per_person": 30000,
        "max_attendees": 10,
        "fixed_first_attendee": "양대호",
    }
    if not INFO_YML.exists():
        return defaults
    rules = read_information_yml().get("attendee_rules", {})
    if isinstance(rules, dict):
        return {**defaults, **rules}
    return defaults


def configured_meeting_places() -> dict[str, str]:
    defaults = {
        "gachon_univ": "바이오나노연구원 315호",
        "seoul_west": "이화여자대학교 의과대학 1109호",
        "ewha_mokdong": "이대목동병원 MCC A 지하 1층 세미나실",
        "kriss": "한국표준과학연구원 313동 1층 회의공간",
    }
    if not INFO_YML.exists():
        return defaults
    places = read_information_yml().get("meeting_places", {})
    if isinstance(places, dict):
        return {**defaults, **{str(key): str(value) for key, value in places.items()}}
    return defaults


def configured_districts(key: str, fallback: tuple[str, ...]) -> tuple[str, ...]:
    if not INFO_YML.exists():
        return fallback
    districts = read_information_yml().get(key, [])
    if isinstance(districts, list):
        cleaned = tuple(str(item) for item in districts if item)
        if cleaned:
            return cleaned
    return fallback


def external_place_keys() -> set[str]:
    if not INFO_YML.exists():
        return {"seoul_west", "ewha_mokdong"}
    config = read_information_yml()
    keys: set[str] = set()
    raw_members = config.get("external_members", [])
    if isinstance(raw_members, list):
        for raw in raw_members:
            if not isinstance(raw, dict):
                continue
            raw_places = raw.get("meeting_places", [])
            if isinstance(raw_places, list):
                keys.update(str(place) for place in raw_places if place)
    return keys or {"seoul_west", "ewha_mokdong"}


def meeting_place_key(meeting_place: str) -> str | None:
    for key, value in configured_meeting_places().items():
        if value == meeting_place:
            return key
    return None


def is_external_meeting_place(meeting_place: str) -> bool:
    key = meeting_place_key(meeting_place)
    return key in external_place_keys() if key else False


def rank_to_position(rank: object) -> str:
    try:
        numeric = int(float(rank))
    except (TypeError, ValueError):
        return "연구원"
    return {0: "교수", 1: "연구원", 2: "학생"}.get(numeric, "연구원")


def read_members(path: Path = MEMBERS_XLSX) -> list[Member]:
    if INFO_YML.exists():
        config = read_information_yml()
        raw_members = config.get("members", [])
        if not isinstance(raw_members, list):
            raise ValueError("information.yml field 'members' must be a list")
        members: list[Member] = []
        for raw in raw_members:
            if not isinstance(raw, dict):
                continue
            name = raw.get("name")
            if not name:
                continue
            members.append(
                Member(
                    department=str(raw.get("department") or "물리학과"),
                    position=str(raw.get("position") or "연구원"),
                    name=str(name),
                )
            )
        return members
    if load_workbook is None:
        return [
            Member("물리학과", "교수", "양대호"),
            Member("물리학과", "연구원", "황희성"),
            Member("물리학과", "연구원", "송영빈"),
            Member("물리학과", "연구원", "장하나"),
            Member("물리학과", "연구원", "Balakiruthika Periyasamy"),
            Member("물리학과", "연구원", "Tayyab Raza"),
            Member("물리학과", "학생", "김예진"),
            Member("물리학과", "학생", "김효현"),
            Member("물리학과", "학생", "김선우"),
            Member("물리학과", "학생", "이정우"),
        ]
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    members: list[Member] = []
    for name, rank in sheet.iter_rows(min_row=2, values_only=True):
        if not name:
            continue
        members.append(Member("물리학과", rank_to_position(rank), str(name)))
    return members


def read_external_members(meeting_place: str) -> list[Member]:
    place_key = meeting_place_key(meeting_place)
    if not place_key or not INFO_YML.exists():
        return []
    config = read_information_yml()
    raw_members = config.get("external_members", [])
    if not isinstance(raw_members, list):
        return []
    members: list[Member] = []
    for raw in raw_members:
        if not isinstance(raw, dict):
            continue
        raw_places = raw.get("meeting_places", [])
        if isinstance(raw_places, list) and place_key not in [str(place) for place in raw_places]:
            continue
        name = raw.get("name")
        if not name:
            continue
        members.append(
            Member(
                department=str(raw.get("department") or ""),
                position=str(raw.get("position") or ""),
                name=str(name),
            )
        )
    return members


def numbered_lines(lines: Iterable[str]) -> str:
    return "\n".join(f"{idx}. {line}" for idx, line in enumerate(lines, 1))


def update_summary(
    metadata: ReceiptMetadata,
    meeting_place: str,
    attendee_total: int,
    attendees: list[Member],
) -> None:
    SUMMARY_CSV.parent.mkdir(parents=True, exist_ok=True)
    records = read_summary_records()
    new_record = {
        "file_name": metadata.receipt_name,
        "total_price": str(metadata.total_price),
        "store_name": metadata.store_name,
        "address": metadata.address,
        "meeting_place": meeting_place,
        "generated": metadata.generated.strftime("%Y-%m-%d %H:%M:%S"),
        "topic": metadata.topic,
        "attendee_count": str(attendee_total),
        "item_count": "" if metadata.item_count is None else str(metadata.item_count),
        "food_count": "" if metadata.food_count is None else str(metadata.food_count),
        "drink_count": "" if metadata.drink_count is None else str(metadata.drink_count),
        "attendee_names": ";".join(member.name for member in attendees),
        "ocr_engine": metadata.ocr_engine,
    }

    for idx, record in enumerate(records):
        if record.get("file_name") == metadata.receipt_name:
            records[idx] = {**record, **new_record}
            break
    else:
        records.append(new_record)

    with SUMMARY_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=SUMMARY_HEADER, extrasaction="ignore")
        writer.writeheader()
        for record in records:
            writer.writerow({column: record.get(column, "") for column in SUMMARY_HEADER})


def register_korean_fonts() -> str:
    for font_name in ("HYSMyeongJo-Medium", "HYGothic-Medium"):
        try:
            pdfmetrics.registerFont(UnicodeCIDFont(font_name))
        except Exception:
            pass
    return "HYGothic-Medium"


def field_rectangles(template_pdf: Path) -> dict[str, list[float]]:
    reader = PdfReader(str(template_pdf))
    page = reader.pages[0]
    rects: dict[str, list[float]] = {}
    for annot_ref in page.get("/Annots") or []:
        annot = annot_ref.get_object()
        name = annot.get("/T")
        if name:
            rects[str(name)] = [float(value) for value in annot.get("/Rect")]
    return rects


def draw_field(
    pdf_canvas: canvas.Canvas,
    rects: dict[str, list[float]],
    font: str,
    name: str,
    text: str,
    size: float = 11,
    leading: float | None = None,
    xpad: float = 4,
    yoffset: float = -3,
) -> None:
    if name not in rects:
        return
    x1, y1, _x2, y2 = rects[name]
    pdf_canvas.setFont(font, size)
    if "\n" not in text:
        y = (y1 + y2) / 2 - size / 2 + 2 + yoffset
        pdf_canvas.drawString(x1 + xpad, y, text)
        return

    text_object = pdf_canvas.beginText(x1 + xpad, y2 - size - 6)
    text_object.setFont(font, size)
    text_object.setLeading(leading or size + 4)
    for line in text.split("\n"):
        text_object.textLine(line)
    pdf_canvas.drawText(text_object)


def name_font_size(name: str) -> float:
    if len(name) > 18:
        return 4.5
    if len(name) > 12:
        return 8
    return 10


def make_minutes_pdf(
    metadata: ReceiptMetadata,
    project_info: dict[str, str],
    attendees: list[Member],
    meeting_place: str,
    output_pdf: Path,
) -> None:
    font = register_korean_fonts()
    rects = field_rectangles(TEMPLATE_PDF)
    topics = configured_all_topics()
    if metadata.topic not in topics:
        raise ValueError(f"Unknown topic '{metadata.topic}'. Check information.yml topics/topic_order.")
    title, body_lines = topics[metadata.topic]

    packet = BytesIO()
    pdf_canvas = canvas.Canvas(packet, pagesize=A4)

    fields = {
        "과제번호": project_info["과제번호"],
        "연구책임자": project_info["연구책임자"],
        "연구과제명": project_info["연구과제명"],
        "연구기간": project_info["연구기간"],
        "회의제목": title,
        "회의일시": metadata.generated.strftime("%Y년 %m월 %d일 %H:%M"),
        "회의장소": meeting_place,
        "회의내용": numbered_lines(body_lines),
    }
    for key in ("과제번호", "연구책임자", "연구과제명", "연구기간", "회의제목", "회의일시", "회의장소"):
        draw_field(pdf_canvas, rects, font, key, fields[key], 11)
    draw_field(pdf_canvas, rects, font, "회의내용", fields["회의내용"], 11, leading=18)

    for idx, member in enumerate(attendees, 1):
        draw_field(pdf_canvas, rects, font, f"소속{idx}", member.department, 7, xpad=2, yoffset=-1)
        draw_field(pdf_canvas, rects, font, f"직위{idx}", member.position, 8, xpad=7, yoffset=-1)
        draw_field(
            pdf_canvas,
            rects,
            font,
            f"성명{idx}",
            member.name,
            name_font_size(member.name),
            xpad=3,
            yoffset=-1,
        )

    pdf_canvas.save()
    packet.seek(0)

    reader = PdfReader(str(TEMPLATE_PDF))
    page = reader.pages[0]
    page.merge_page(PdfReader(packet).pages[0])
    if "/Annots" in page:
        del page["/Annots"]

    writer = PdfWriter()
    writer.add_page(page)
    output_pdf.parent.mkdir(parents=True, exist_ok=True)
    with output_pdf.open("wb") as handle:
        writer.write(handle)


def receipt_image_page(receipt_path: Path) -> BytesIO:
    packet = BytesIO()
    pdf_canvas = canvas.Canvas(packet, pagesize=A4)
    image = ImageOps.exif_transpose(Image.open(receipt_path)).convert("RGB")
    max_width, max_height = A4[0] - 56, A4[1] - 56
    scale = min(max_width / image.width, max_height / image.height)
    width, height = image.width * scale, image.height * scale
    pdf_canvas.drawImage(
        ImageReader(image),
        (A4[0] - width) / 2,
        (A4[1] - height) / 2,
        width=width,
        height=height,
        preserveAspectRatio=True,
        mask="auto",
    )
    pdf_canvas.save()
    packet.seek(0)
    return packet


def combine_minutes_and_receipt(minutes_pdf: Path, receipt_path: Path, output_pdf: Path) -> None:
    writer = PdfWriter()
    for page in PdfReader(str(minutes_pdf)).pages:
        writer.add_page(page)
    for page in PdfReader(receipt_image_page(receipt_path)).pages:
        writer.add_page(page)
    with output_pdf.open("wb") as handle:
        writer.write(handle)


def main() -> None:
    args = parse_args()
    metadata = metadata_from_args(args)
    meeting_place = derive_meeting_place(metadata.address, metadata.store_name)
    external_meeting = is_external_meeting_place(meeting_place)
    if metadata.topic == "auto":
        metadata = ReceiptMetadata(
            receipt_path=metadata.receipt_path,
            total_price=metadata.total_price,
            generated=metadata.generated,
            store_name=metadata.store_name,
            address=metadata.address,
            topic=choose_auto_topic(metadata.receipt_name, external=external_meeting),
            item_count=metadata.item_count,
            food_count=metadata.food_count,
            drink_count=metadata.drink_count,
            ocr_engine=metadata.ocr_engine,
        )
    attendee_total = adjusted_attendee_count(
        metadata.total_price,
        metadata.item_count,
        metadata.food_count,
        metadata.drink_count,
    )
    project_info = read_project_info()
    members = read_members()
    external_members = read_external_members(meeting_place) if external_meeting else []
    attendees = select_attendees(members, attendee_total, metadata, external_members=external_members)
    attendee_total = len(attendees)

    minutes_pdf = OUTPUT_DIR / f"{metadata.stem}_바나연회의록.pdf"
    combined_pdf = OUTPUT_DIR / f"{metadata.stem}_바나연회의록_영수증첨부.pdf"

    if not args.no_summary:
        update_summary(metadata, meeting_place, attendee_total, attendees)
    make_minutes_pdf(metadata, project_info, attendees, meeting_place, minutes_pdf)
    combine_minutes_and_receipt(minutes_pdf, metadata.receipt_path, combined_pdf)

    print(f"summary_csv: {SUMMARY_CSV}")
    print(f"total_price: {metadata.total_price}")
    print(f"generated: {metadata.generated:%Y-%m-%d %H:%M:%S}")
    print(f"store_name: {metadata.store_name}")
    print(f"address: {metadata.address}")
    print(f"topic: {metadata.topic}")
    print(f"item_count: {metadata.item_count if metadata.item_count is not None else ''}")
    print(f"food_count: {metadata.food_count if metadata.food_count is not None else ''}")
    print(f"drink_count: {metadata.drink_count if metadata.drink_count is not None else ''}")
    print(f"attendee_count: {attendee_total}")
    print(f"attendee_names: {';'.join(member.name for member in attendees)}")
    print(f"meeting_place: {meeting_place}")
    print(f"minutes_pdf: {minutes_pdf}")
    print(f"combined_pdf: {combined_pdf}")


if __name__ == "__main__":
    main()
